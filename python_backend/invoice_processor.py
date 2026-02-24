import ollama
import base64
import os
from pathlib import Path
import json
import openpyxl
import shutil

# Get the directory where this script is located
SCRIPT_DIR = Path(__file__).parent.absolute()
INPUT_FOLDER = SCRIPT_DIR / "temp_input"
OUTPUT_FOLDER = SCRIPT_DIR / "output"

# Ensure directories exist
INPUT_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)


# --- Agent 1: OCRAgent ---
class OCRAgent:
    """Uses a multimodal model to extract structured entities from an image."""

    def __init__(self, model='qwen2.5vl:7b'):
        self.model = model
        print(f"✅ OCR Agent ready, using model: {self.model}")

    def _encode_image(self, image_path):
        """Encodes an image to base64."""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')

    def execute(self, image_path: Path):
        """Takes an image path and returns a simplified dictionary."""
        print(f"🔬 OCR Agent: Analyzing image '{image_path.name}'...")
        response_text = ""
        try:
            encoded_image = self._encode_image(image_path)
            prompt = """
            You are an expert invoice extraction AI.
            Extract only the following fields in a valid JSON:
            {
              "Seller": {"Name": "", "GSTIN": ""},
              "Buyer": {"GSTIN": ""},
              "Invoice": {
                "BillDate": "",
                "Items": [
                  {"SNo": 1, "HSNCode": "", "CGST_Percent": 0, "CGST": 0, "SGST_Percent": 0, "SGST": 0, "Total": 0}
                ],
                "NetPayableAmount": 0
              }
            }
            Fill with exact values from the invoice image.
            """
            response = ollama.chat(
                model=self.model,
                messages=[{'role': 'user', 'content': prompt, 'images': [encoded_image]}],
                options={'temperature': 0.0}
            )
            response_text = response['message']['content']

            json_start = response_text.find('{')
            json_end = response_text.rfind('}')
            if json_start != -1 and json_end != -1:
                json_string = response_text[json_start: json_end + 1]
                return json.loads(json_string)
            else:
                raise json.JSONDecodeError("No JSON object found in model response", response_text, 0)
        except Exception as e:
            print(f"❌ OCR Agent Error: {e}")
            if response_text: print(f"   Raw response was: {response_text}")
            return None


# --- Agent 2: ExcelAgent ---
class ExcelAgent:
    """Creates simplified GST-formatted Excel file."""

    def __init__(self, model='deepseek-coder:6.7b'):
        self.model = model
        print(f"✅ Excel Agent ready, using model: {self.model}")

    def execute(self, task_data: dict):
        invoice_data = task_data.get('invoice_data')
        filename = task_data.get('filename')

        if not all([invoice_data, filename]):
            print("❌ Excel Agent Error: Missing 'invoice_data' or 'filename'.")
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GST Report"

            # --- Top Info ---
            ws['A1'] = "Seller GST:"
            ws['B1'] = invoice_data.get("Seller", {}).get("GSTIN", "")
            ws['A2'] = "Buyer GST:"
            ws['B2'] = invoice_data.get("Buyer", {}).get("GSTIN", "")
            ws['A3'] = "Bill Date:"
            ws['B3'] = invoice_data.get("Invoice", {}).get("BillDate", "")
            ws['A4'] = "Seller Name:"
            ws['B4'] = invoice_data.get("Seller", {}).get("Name", "")

            # --- Table Header ---
            headers = ["S.No", "HSN Code", "CGST %", "CGST Amount", "SGST %", "SGST Amount", "Total Value"]
            ws.append([])
            ws.append(headers)

            # --- Items Rows ---
            for item in invoice_data.get("Invoice", {}).get("Items", []):
                row_data = [
                    item.get("SNo"),
                    item.get("HSNCode"),
                    item.get("CGST_Percent"),
                    item.get("CGST"),
                    item.get("SGST_Percent"),
                    item.get("SGST"),
                    item.get("Total")
                ]
                ws.append(row_data)

            # --- Footer ---
            ws.append([])
            ws.append(
                ["", "", "", "", "", "Net Payable Amount:", invoice_data.get("Invoice", {}).get("NetPayableAmount", 0)])

            wb.save(filename)
            print(f"✅ Excel Agent: Created simplified GST report '{filename}'")
            return True
        except Exception as e:
            print(f"❌ Excel Agent Error: {e}")
            return False


# --- Manager ---
class Manager:
    def __init__(self, agents: dict):
        self.agents = agents
        print(f"🚀 Manager ready with agents: {list(self.agents.keys())}")

    def create_ocr_to_excel_plan(self, image_path: Path):
        # Use the globally defined OUTPUT_FOLDER
        excel_filename = str(OUTPUT_FOLDER / f"{image_path.stem}_gst_report.xlsx")
        return [
            {"task_name": "Extract Entities", "agent_id": "ocr_agent", "input": image_path, "output_key": "ocr_data"},
            {"task_name": "Generate Excel", "agent_id": "excel_agent",
             "input_mapping": {"invoice_data": "ocr_data", "filename": excel_filename}, "output_key": "report_success"}
        ]

    def execute_plan(self, plan: list):
        context = {}
        print("\n--- Executing Plan ---")
        for step in plan:
            agent_id = step['agent_id']
            task_input = step.get('input')
            if 'input_mapping' in step:
                task_input = {k: context.get(v, v) for k, v in step['input_mapping'].items()}
            agent = self.agents.get(agent_id)
            if not agent:
                print(f"❌ Manager Error: Agent '{agent_id}' not found.");
                return None
            print(f"🏃 Task: '{step['task_name']}' -> Agent: '{agent_id}'")
            result = agent.execute(task_input)
            if not result:
                print(f"❌ Task '{step['task_name']}' failed. Aborting.");
                return None
            if 'output_key' in step:
                context[step['output_key']] = result
        print("--- Plan Finished ---")
        return context


def run_workflow(image_path: Path):
    """Run the OCR to Excel workflow for a single image."""
    try:
        # Initialize agents
        ocr_agent = OCRAgent()
        excel_agent = ExcelAgent()
        manager = Manager({"ocr_agent": ocr_agent, "excel_agent": excel_agent})
        
        # Create and execute plan
        plan = manager.create_ocr_to_excel_plan(image_path)
        result = manager.execute_plan(plan)
        
        if result and "report_success" in result and result["report_success"]:
            return {"success": True, "data": result.get("ocr_data")}
        else:
            return {"success": False, "data": None}
    except Exception as e:
        print(f"❌ Workflow Error: {e}")
        return {"success": False, "data": None}


def generate_excel(invoice_data: dict, output_path: Path):
    """Generate Excel file from invoice data."""
    try:
        excel_agent = ExcelAgent()
        success = excel_agent.execute({
            "invoice_data": invoice_data,
            "filename": str(output_path)
        })
        return success
    except Exception as e:
        print(f"❌ Excel Generation Error: {e}")
        return False


def consolidate_excel_files(file_paths: list, output_path: Path):
    """Consolidate multiple Excel files into one."""
    try:
        if not file_paths:
            print("❌ No files to consolidate")
            return False
            
        # Create a new workbook for consolidated data
        consolidated_wb = openpyxl.Workbook()
        consolidated_ws = consolidated_wb.active
        consolidated_ws.title = "Consolidated GST Reports"
        
        # Copy headers from the first file
        if file_paths:
            first_wb = openpyxl.load_workbook(file_paths[0])
            first_ws = first_wb.active
            
            # Copy all rows from the first sheet
            for row in first_ws.iter_rows(values_only=True):
                consolidated_ws.append(row)
                
            # Add data from other files (skip headers)
            for file_path in file_paths[1:]:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Skip the first few rows that contain headers
                # (assuming first 6 rows are headers based on the ExcelAgent format)
                for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_num > 6:  # Skip header rows
                        consolidated_ws.append(row)
                        
            consolidated_wb.save(output_path)
            print(f"✅ Consolidated Excel file created: {output_path}")
            return True
        else:
            print("❌ No files to consolidate")
            return False
    except Exception as e:
        print(f"❌ Consolidation Error: {e}")
        return False


def process_invoice_files(image_paths):
    """Main processing function for multiple invoice images."""
    try:
        print(f"📁 Processing {len(image_paths)} invoice files...")
        print(f"📂 Input folder: {INPUT_FOLDER}")
        print(f"📂 Output folder: {OUTPUT_FOLDER}")
        
        # Ensure output directory exists
        OUTPUT_FOLDER.mkdir(exist_ok=True)
        
        # Copy files to input folder
        images_to_process = []
        for image_path in image_paths:
            src_path = Path(image_path)
            if src_path.exists():
                dest_path = INPUT_FOLDER / src_path.name
                shutil.copy2(src_path, dest_path)
                images_to_process.append(dest_path)
                print(f"📋 Copied {src_path.name} to input folder")
            else:
                print(f"❌ Source file not found: {image_path}")

        processed_count = 0
        generated_reports = []
        
        # Process each image
        for image_path in images_to_process:
            print(f"\n--- Processing {image_path.name} ---")
            
            # Run the workflow
            workflow_result = run_workflow(image_path)
            
            if workflow_result["success"]:
                # Generate Excel file
                excel_filename = f"{image_path.stem}_gst_report.xlsx"
                dest_path = OUTPUT_FOLDER / excel_filename
                
                print(f"📊 Generating Excel: {excel_filename}")
                success = generate_excel(workflow_result["data"], dest_path)
                
                if success and dest_path.exists():
                    print(f"✅ Generated Excel file: {dest_path}")
                    
                    # Add to generated reports list with absolute path
                    generated_reports.append({
                        "filename": excel_filename,
                        "file_path": str(dest_path.absolute()),  # Use absolute path
                        "created_at": os.path.getctime(dest_path)
                    })
                    print(f"✅ Added report to generated reports: {excel_filename}")
                    processed_count += 1
                else:
                    print(f"❌ Error: Excel file '{dest_path}' not found.")
            else:
                print(f"❌ Workflow failed for '{image_path.name}'")

        print(f"📊 Processed {processed_count} out of {len(images_to_process)} files successfully")
        print(f"📄 Generated reports: {len(generated_reports)}")
        
        # List all files in output directory for debugging
        print("📁 Files in output directory:")
        for file in OUTPUT_FOLDER.iterdir():
            if file.is_file():
                print(f"  - {file.name} ({file.stat().st_size} bytes)")

        # Create consolidated output
        consolidated_output = OUTPUT_FOLDER / "Consolidated_Invoices_Output.xlsx"
        if generated_reports:
            consolidate_excel_files([report["file_path"] for report in generated_reports], consolidated_output)
            print(f"📊 Consolidated output created: {consolidated_output}")
        else:
            print("❌ No reports to consolidate")

        if len(generated_reports) > 0:
            result = {
                "success": True,
                "message": f"Successfully processed {len(generated_reports)} out of {len(images_to_process)} files",
                "output_file": str(consolidated_output.absolute()),  # Use absolute path
                "invoices_count": len(generated_reports),
                "generated_reports": generated_reports  # Include generated reports info
            }
            print(f"[RESULT] {json.dumps(result)}")
            return result
        else:
            result = {
                "success": False,
                "message": "No files were successfully processed",
                "output_file": None,
                "invoices_count": 0,
                "generated_reports": []
            }
            print(f"[RESULT] {json.dumps(result)}")
            return result

    except Exception as e:
        print(f"❌ Processing Error: {e}")
        import traceback
        traceback.print_exc()
        result = {
            "success": False,
            "message": f"Processing failed: {str(e)}",
            "output_file": None,
            "invoices_count": 0,
            "generated_reports": []
        }
        print(f"[RESULT] {json.dumps(result)}")
        return result
