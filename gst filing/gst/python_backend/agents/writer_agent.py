from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def writer_agent(data: list, file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "GST_Invoices"

    bold_font = Font(bold=True)
    
    # Headers matching your required format exactly
    headers = [
        "S.No.", "Vendor/Shop Name", "Date", "GSTIN", "Invoice No.", 
        "HSN Codes", "CGST", "SGST", "IGST", "Total Tax", "Taxable Amount"
    ]

    # Create border styles for clean look
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header styling - bold without colors
    header_font = Font(bold=True, size=12)
    
    # Set enhanced column widths and create headers
    for col, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Set enhanced column widths for better readability
        if header == "S.No.":
            ws.column_dimensions[column_letter].width = 12
        elif header == "Vendor/Shop Name":
            ws.column_dimensions[column_letter].width = 35
        elif header == "Date":
            ws.column_dimensions[column_letter].width = 18
        elif header == "GSTIN":
            ws.column_dimensions[column_letter].width = 20
        elif header == "Invoice No.":
            ws.column_dimensions[column_letter].width = 20
        elif header == "HSN Codes":
            ws.column_dimensions[column_letter].width = 35
        elif header in ["CGST", "SGST", "IGST"]:
            ws.column_dimensions[column_letter].width = 15
        elif header == "Total Tax":
            ws.column_dimensions[column_letter].width = 18
        elif header == "Taxable Amount":
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 18

    # Process data rows
    for row_idx, record in enumerate(data, start=2):
        # S.No.
        cell = ws.cell(row=row_idx, column=1)
        cell.value = row_idx - 1
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Vendor/Shop Name
        shop_name = record.get("Shop Name", "")
        if shop_name in (None, "", "null", 0, 0.0):
            shop_name = "N/A"
        cell = ws.cell(row=row_idx, column=2)
        cell.value = shop_name
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        cell.border = thin_border
        
        # Date (Invoice Date)
        date_value = record.get("Invoice Date", "")
        if date_value in (None, "", "null", 0, 0.0):
            date_value = "N/A"
        cell = ws.cell(row=row_idx, column=3)
        cell.value = date_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # GSTIN
        gstin_value = record.get("GSTIN", "")
        if gstin_value in (None, "", "null", 0, 0.0):
            gstin_value = "N/A"
        cell = ws.cell(row=row_idx, column=4)
        cell.value = gstin_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Invoice No.
        invoice_value = record.get("Invoice Number", "")
        if invoice_value in (None, "", "null", 0, 0.0):
            invoice_value = "N/A"
        cell = ws.cell(row=row_idx, column=5)
        cell.value = invoice_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # HSN Codes - Format with 3 codes per line for better readability
        hsn_value = record.get("HSN Code", "")
        if hsn_value in (None, "", "null", 0, 0.0):
            hsn_value = "N/A"
        elif isinstance(hsn_value, list):
            # Clean valid codes
            valid_codes = [str(x).strip() for x in hsn_value if str(x).strip() not in ("", "0", "null")]
            if valid_codes:
                # Group codes into chunks of 3 per line
                lines = []
                for i in range(0, len(valid_codes), 3):
                    chunk = valid_codes[i:i+3]
                    lines.append(", ".join(chunk))
                hsn_value = "\n".join(lines)
            else:
                hsn_value = "N/A"
        elif isinstance(hsn_value, str) and "," in hsn_value:
            # Clean up comma-separated codes and format with 3 per line
            valid_codes = [x.strip() for x in hsn_value.split(",") if x.strip() not in ("", "0", "null")]
            if valid_codes:
                # Group codes into chunks of 3 per line
                lines = []
                for i in range(0, len(valid_codes), 3):
                    chunk = valid_codes[i:i+3]
                    lines.append(", ".join(chunk))
                hsn_value = "\n".join(lines)
            else:
                hsn_value = "N/A"
        
        cell = ws.cell(row=row_idx, column=6)
        cell.value = hsn_value
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = thin_border
        
        # CGST
        cgst_value = record.get("CGST", "")
        if cgst_value in (None, "", "null", 0, 0.0):
            cgst_value = "N/A"
        cell = ws.cell(row=row_idx, column=7)
        cell.value = cgst_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # SGST
        sgst_value = record.get("SGST", "")
        if sgst_value in (None, "", "null", 0, 0.0):
            sgst_value = "N/A"
        cell = ws.cell(row=row_idx, column=8)
        cell.value = sgst_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # IGST
        igst_value = record.get("IGST", "")
        if igst_value in (None, "", "null", 0, 0.0):
            igst_value = "N/A"
        cell = ws.cell(row=row_idx, column=9)
        cell.value = igst_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Total Tax
        total_tax = record.get("Tax Amount", "")
        if total_tax in (None, "", "null", 0, 0.0):
            # Try to calculate from individual taxes if available
            try:
                cgst = float(cgst_value) if cgst_value not in ("N/A", None, "") else 0
                sgst = float(sgst_value) if sgst_value not in ("N/A", None, "") else 0
                igst = float(igst_value) if igst_value not in ("N/A", None, "") else 0
                calculated_total = cgst + sgst + igst
                total_tax = calculated_total if calculated_total > 0 else "N/A"
            except (ValueError, TypeError):
                total_tax = "N/A"
        cell = ws.cell(row=row_idx, column=10)
        cell.value = total_tax
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Taxable Amount
        taxable_value = record.get("Total Amount", "")
        if taxable_value in (None, "", "null", 0, 0.0):
            taxable_value = "N/A"
        cell = ws.cell(row=row_idx, column=11)
        cell.value = taxable_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Set enhanced row height - adjust for HSN codes with multiple lines
        if "\n" in str(hsn_value) and hsn_value != "N/A":
            # Count lines in HSN codes and adjust row height accordingly
            line_count = len(str(hsn_value).split("\n"))
            # Each line needs about 18-20 pixels, with some padding
            ws.row_dimensions[row_idx].height = max(40, 20 * line_count + 15)
        else:
            ws.row_dimensions[row_idx].height = 40

    # Set enhanced header row height
    ws.row_dimensions[1].height = 40
    
    # Freeze the header row for better navigation
    ws.freeze_panes = 'A2'
    
    # Keep gridlines visible for better readability
    ws.sheet_view.showGridLines = True

    wb.save(file_path)
    print(f"[Writer Agent] Excel report saved to: {file_path}")